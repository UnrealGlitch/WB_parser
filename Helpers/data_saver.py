import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if __name__ == "__main__":
    raise Exception("Нужно запускать wb_parser.py")

class DataSaver:
    
    def __init__(self):
        pass

    def make_response_file(self, response_data):
        with open("Data/input.txt", "w", encoding="utf-8") as file:
            json.dump(response_data, file, ensure_ascii=False, indent=2)
    
    def save_xlsx(self, rows: list[dict], output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары WB"

        headers = list(rows[0].keys())

        header_fill = PatternFill("solid", start_color="4472C4")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        data_font = Font(name="Arial", size=10)
        alt_fill = PatternFill("solid", start_color="DCE6F1")

        for i in range(0, len(rows)):
            record = rows[i]
            for j in range(0, len(headers)):
                key = headers[j]
                cell = ws.cell(row=i + 2, column=j + 1, value=record[key])
                cell.font = data_font
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        col_widths = {
            "Ссылка на товар":       60,
            "Артикул":               14,
            "Название":              35,
            "Цена (руб)":            12,
            "Описание":              30,
            "Ссылки на изображения": 60,
            "Бренд":                 20,
            "Тип товара":            14,
            "Цвет":                  18,
            "Рейтинг":               10,
            "Количество отзывов":    18,
            "Размеры":               25,
            "Остатки":               10,
            "Название селлера":      25,
            "Ссылка на селлера":     35,
            "Рейтинг селлера":       16
        }
        for i in range(0, len(headers)):
            ws.column_dimensions[get_column_letter(i + 1)].width = col_widths.get(headers[i], 20)

        ws.freeze_panes = "A2"
        wb.save(output_path)